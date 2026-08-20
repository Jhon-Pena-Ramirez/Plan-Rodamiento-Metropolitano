"""
Lectura y normalizacion de los 3 archivos de entrada:
- Base de vehiculos
- Horarios (topes y recorridos por ruta fisica)
- Sorteo (orden inicial de cada corredor)
"""
import openpyxl
from datetime import time as dtime

from config_corredores import HOJA_HORARIOS, GRUPO_SORTEO, GRUPO_NUEVA_FLOTA


def _to_hhmm(value):
    """Normaliza un valor de celda de hora a string 'HH:MM'."""
    if value is None:
        return None
    if isinstance(value, dtime):
        return value.strftime("%H:%M")
    if hasattr(value, "strftime"):  # datetime
        return value.strftime("%H:%M")
    s = str(value).strip()
    # admite '03:35:00' o '03:35'
    parts = s.split(":")
    if len(parts) >= 2:
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return s


# ---------------------------------------------------------------------------
# VEHICULOS
# ---------------------------------------------------------------------------
def load_vehiculos(path):
    """
    Retorna dict: veh_interno (str, sin ceros a la izquierda) -> {
        interno, placa, nui, tarjeta_operacion, tipo
    }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    vehiculos = {}
    header_row = None
    for row in range(1, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if vals and any(v and "interno" in str(v).lower() for v in vals):
            header_row = row
            break
    if header_row is None:
        raise ValueError("No se encontro la fila de encabezados en Base_vehiculos.xlsx")

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    col_interno = headers.get("veh_interno")
    col_placa = headers.get("veh_placa")
    col_nui = headers.get("nui")
    col_tarjeta = headers.get("veh_tarjeta_operacion")
    col_tipo = headers.get("veh_tipo")

    for row in range(header_row + 1, ws.max_row + 1):
        interno = ws.cell(row=row, column=col_interno).value
        if interno is None or str(interno).strip() == "":
            continue
        interno_norm = str(interno).strip().lstrip("0") or "0"
        vehiculos[interno_norm] = {
            "interno": str(interno).strip(),
            "placa": ws.cell(row=row, column=col_placa).value,
            "nui": ws.cell(row=row, column=col_nui).value,
            "tarjeta_operacion": ws.cell(row=row, column=col_tarjeta).value,
            "tipo": ws.cell(row=row, column=col_tipo).value,
        }
    return vehiculos


# ---------------------------------------------------------------------------
# HORARIOS
# ---------------------------------------------------------------------------
def _parse_tabla_recorridos(ws, fila_titulo):
    """
    A partir de la fila donde esta el titulo 'RECORRIDOS ...', parsea:
    - fila_titulo:      NaN | 'RECORRIDOS...' | 'NOMBRE RUTA' | ...
    - fila_titulo + 1:  NaN | NaN | 1 | 2 | 3 | 4 ...   (numero de bloque)
    - filas siguientes: NaN | posicion | hora1 | hora2 | ...
    Retorna: tope (int), bloques (list[int]), matrix (dict pos -> {bloque: 'HH:MM'})
    """
    fila_bloques = fila_titulo + 1
    bloques = []
    col = 3  # columna C en adelante
    while True:
        v = ws.cell(row=fila_bloques, column=col).value
        if v is None:
            break
        bloques.append(int(v))
        col += 1
    n_bloques = len(bloques)

    matrix = {}
    r = fila_titulo + 2
    pos = 0
    while True:
        v_pos = ws.cell(row=r, column=2).value
        if v_pos is None:
            break
        pos += 1
        fila_horas = {}
        for i, bloque in enumerate(bloques):
            hora = ws.cell(row=r, column=3 + i).value
            fila_horas[bloque] = _to_hhmm(hora)
        matrix[pos] = fila_horas
        r += 1

    return pos, bloques, matrix  # pos terminó siendo el tope (cantidad de filas)


def load_horarios(path):
    """
    Retorna dict: ruta_fisica -> {
        'habil':   {'tope': N, 'bloques': [...], 'matrix': {pos: {bloque: 'HH:MM'}}},
        'festivo': {...}
    }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    resultado = {}

    for ruta_fisica, nombre_hoja in HOJA_HORARIOS.items():
        if nombre_hoja not in wb.sheetnames:
            raise ValueError(f"No se encontro la hoja '{nombre_hoja}' en Horarios")
        ws = wb[nombre_hoja]

        titulos = []
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=2).value
            if v and "RECORRIDOS" in str(v).upper():
                titulos.append(row)
        if len(titulos) < 2:
            raise ValueError(f"La hoja '{nombre_hoja}' no tiene las 2 tablas esperadas "
                              f"(normal y dominical/festivo)")

        tope_h, bloques_h, matrix_h = _parse_tabla_recorridos(ws, titulos[0])
        tope_f, bloques_f, matrix_f = _parse_tabla_recorridos(ws, titulos[1])

        nombre_real = ws.cell(row=titulos[0], column=3).value
        nombre_real = str(nombre_real).strip() if nombre_real else ruta_fisica

        resultado[ruta_fisica] = {
            "nombre_real": nombre_real,
            "habil": {"tope": tope_h, "bloques": bloques_h, "matrix": matrix_h},
            "festivo": {"tope": tope_f, "bloques": bloques_f, "matrix": matrix_f},
        }
    return resultado


# ---------------------------------------------------------------------------
# SORTEO
# ---------------------------------------------------------------------------
def load_sorteo(path):
    """
    Retorna dict: corredor -> [interno1, interno2, ...] (orden = posicion inicial)
    Ademas incluye 'NUEVA_FLOTA' -> [...]
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # localizar fila de titulos de grupo (fila 1) y fila de subencabezados (INTERNO/PLACA/NUI)
    fila_grupos = None
    for row in range(1, ws.max_row + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v and "INTERNO" in str(v).upper() for v in vals):
            fila_grupos = row - 1
            fila_sub = row
            break
    if fila_grupos is None:
        raise ValueError("No se encontro la fila de encabezados en Sorteo.xlsx")

    # localizar columna 'INTERNO' de cada grupo
    columnas_interno = {}
    grupo_actual = None
    for c in range(1, ws.max_column + 1):
        v_grupo = ws.cell(row=fila_grupos, column=c).value
        if v_grupo:
            grupo_actual = str(v_grupo).strip().upper()
        v_sub = ws.cell(row=fila_sub, column=c).value
        if v_sub and "INTERNO" in str(v_sub).upper() and grupo_actual:
            columnas_interno[grupo_actual] = c

    resultado = {}
    nombre_a_clave = {v: k for k, v in GRUPO_SORTEO.items()}

    for nombre_grupo, col in columnas_interno.items():
        lista = []
        for row in range(fila_sub + 1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None or str(v).strip() in ("", "-"):
                continue
            lista.append(str(v).strip().lstrip("0") or "0")

        if nombre_grupo == GRUPO_NUEVA_FLOTA:
            resultado["NUEVA_FLOTA"] = lista
        else:
            clave = nombre_a_clave.get(nombre_grupo)
            if clave:
                resultado[clave] = lista
            else:
                # grupo desconocido: se guarda tal cual para no perder informacion
                resultado[nombre_grupo] = lista

    return resultado
