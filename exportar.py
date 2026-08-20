"""
Genera el archivo .xlsx final: una hoja por ruta fisica, columnas
Ruta, Fecha, Placa, NUI, Hora -- ordenado por fecha y hora, con los
domingos/festivos resaltados.

La Fecha y la Hora se escriben como valores REALES de Excel (no como texto),
para que el formato de celda, el ordenamiento y los filtros funcionen bien
del lado del usuario.
"""
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

FILL_FESTIVO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(name="Arial", bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Arial")


def _hora_a_time(hora_str):
    """Convierte 'HH:MM' a datetime.time real para que Excel la reconozca como hora."""
    if hora_str is None:
        return None
    return datetime.strptime(hora_str, "%H:%M").time()


def exportar_plan(todas_las_filas, vehiculos, horarios, path_salida):
    """
    todas_las_filas: dict ruta_fisica -> lista de dicts
        {'fecha': date, 'hora': 'HH:MM', 'veh_interno': str, 'tipo_dia': str}
    vehiculos: dict de io_data.load_vehiculos()
    horarios: dict de io_data.load_horarios()  (para tomar el nombre_real de cada ruta)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ruta, filas in todas_las_filas.items():
        nombre_real = horarios.get(ruta, {}).get("nombre_real", ruta)
        ws = wb.create_sheet(title=ruta[:31])  # excel limita el nombre de hoja a 31 caracteres

        headers = ["Ruta", "Fecha", "Placa", "NUI", "Hora"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center")

        filas_ordenadas = sorted(filas, key=lambda f: (f["fecha"], f["hora"]))

        r = 2
        for f in filas_ordenadas:
            interno = f["veh_interno"]
            veh = vehiculos.get(interno, {})
            placa = veh.get("placa", f"(desconocido:{interno})")
            nui = veh.get("nui", "")

            ws.cell(row=r, column=1, value=nombre_real)

            c_fecha = ws.cell(row=r, column=2, value=f["fecha"])  # objeto date real
            c_fecha.number_format = "DD/MM/YYYY"

            ws.cell(row=r, column=3, value=placa)
            ws.cell(row=r, column=4, value=nui)

            c_hora = ws.cell(row=r, column=5, value=_hora_a_time(f["hora"]))  # objeto time real
            c_hora.number_format = "HH:MM"

            if f.get("tipo_dia") == "festivo":
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = FILL_FESTIVO
            for c in range(1, 6):
                ws.cell(row=r, column=c).font = FONT_NORMAL
            r += 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 10
        ws.freeze_panes = "A2"

    wb.save(path_salida)
    return path_salida
